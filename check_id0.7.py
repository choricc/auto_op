import concurrent.futures
import time
import requests
import json
from datetime import datetime, timedelta
from gooey import Gooey, GooeyParser
import sys
from openpyxl import Workbook, load_workbook
import os

today = datetime.today()
tomorrow = today + timedelta(days=1)


@Gooey(progress_regex=r"^progress: (\d+)/(\d+)$",
       progress_expr="x[0] / x[1] * 100",
       disable_progress_bar_animation=True,
       language='chinese')
def main():
    parser = GooeyParser(description="云窍数据检查")
    parser.add_argument('--token', type=str, help='f12获取token')
    parser.add_argument('--check', type=str, metavar="物理机/云主机检查", widget="Dropdown",
                        choices=['物理机', '云主机'])
    parser.add_argument('--js', widget="CheckBox", action="store_true", metavar="江苏")
    parser.add_argument('--gs', widget="CheckBox", action="store_true", metavar="甘肃")
    parser.add_argument('--hb', widget="CheckBox", action="store_true", metavar="河北")
    parser.add_argument('--tj', widget="CheckBox", action="store_true", metavar="天津")
    parser.add_argument('--sx', widget="CheckBox", action="store_true", metavar="山西")
    parser.add_argument('--nmg', widget="CheckBox", action="store_true", metavar="内蒙古")
    parser.add_argument('--qh', widget="CheckBox", action="store_true", metavar="青海")

    args = parser.parse_args(sys.argv[1:])
    choose_type = args.check
    CheckPool = [args.js, args.gs, args.hb, args.tj, args.sx, args.nmg, args.qh]
    poolcicds = choose_pool(CheckPool, choose_type)
    get_cicd(args.token, poolcicds, choose_type)
    # print(poolcicd[0])


def choose_pool(poold, choose_type):
    replace_list = []
    poolcicd = []
    dic_pool = {
        '江苏': '6180578038500323',
        '甘肃': '6180578038500313',
        '河北': '6180578038500318',
        '天津': '6180578038500334',
        '山西': '6180578038500333',
        '内蒙古': '6180578038500326',
        '青海': '6180578038500328'
    }
    yun_dic_pool = {
        '甘肃': '6180578038500313',
        '河北': '6180578038500318',
        '江苏': '6180578038500323',
        '内蒙古': '6180578038500326',
        '青海': '6180578038500328',
        '山西': '6180578038500333',
        '天津': '6180578038500334'
    }
    if choose_type == '云主机':
        dic_pool = yun_dic_pool
    print(dic_pool)
    poolist = ['江苏', '甘肃', '河北', '天津', '山西', '内蒙古', '青海']
    for i in range(len(poold)):
        if poold[i]:
            replace_list.append(poolist[i])
            poolcicd.append(dic_pool.get(poolist[i]))
    return poolcicd


def get_cicd(token, poolcicds, choose_type):
    url = 'http://172.16.21.225/portal/monitorAxios/device/adapter/getDevices'
    headers = {

        'content-type': 'application/json;charset=UTF-8',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.67',
        'token': token
    }
    if choose_type == '云主机':
        for poolcicd in poolcicds:
            count = 0
            dic = {}
            for num in range(20):
                json_data = {
                    "page":
                        {
                            'current': num + 1,
                            'size': '500',
                            'total': '226188'
                        },
                    "query":
                        {
                            'agentVersion': "",
                            'areaCiCode': "6170301785850243",
                            'devName': "",
                            'devType': "VmServer",
                            'devTypeGroup': "云资源",
                            'deviceStatus': "",
                            'exactIp': 'false',
                            'hostServerIp': 'null',
                            'ipAddress': "",
                            'podCiCode': "",
                            'regionCiCode': "",
                            'resourcePoolCiCode': poolcicd,
                            'serverDimensionState': "",
                            'standardBusinessName': "",
                            'tenantName': "",
                            'vendor': ""
                        }
                }
                response = requests.post(url=url, headers=headers, json=json_data)
                # print(response.text)
                json_data = json.loads(response.text)
                for i in range(len(json_data['data']['records'])):
                    count += 1
                    key = json_data['data']['records'][i]['resourceId']
                    value = json_data['data']['records'][i]['name']
                    dic[key] = value
                print(count)
                check_id(dic, token, poolcicd, count)
    else:
        for poolcicd in poolcicds:
            count = 0
            dic = {}
            for num in range(20):
                json_data = {
                    "page":
                        {
                            'current': num + 1,
                            'size': '500',
                            'total': '388894'
                        },
                    "query":
                        {
                            'areaCiCode': "6170301785850243",
                            'devType': "Server",
                            'deviceDescribe': "KVM宿主机,裸金属服务器Linux,VMware宿主机,裸金属legacy,裸金属服务器Windows",
                            'deviceStatus': "运行",
                            'exactIp': 'false',
                            'resourcePoolCiCode': poolcicd
                        }
                }
                response = requests.post(url=url, headers=headers, json=json_data)
                # print(response.text)
                json_data = json.loads(response.text)
                for i in range(len(json_data['data']['records'])):
                    count += 1
                    key = json_data['data']['records'][i]['resourceId']
                    value = json_data['data']['records'][i]['ciCode']
                    dic[key] = value
                print(count)
                check_id(dic, token, poolcicd, count)
        # return dic, count


def check_id(dic_cicd, token, poolcicd, numbers, choose_type):
    # ... （保留原有变量初始化和文件打开部分）...
    b_time = today.strftime('%Y-%m-%d') + ' 00:00:00'
    e_time = tomorrow.strftime('%Y-%m-%d') + ' 00:00:00'
    H = today.strftime('%H')
    count = int(H) * 12
    url = 'http://172.16.21.225/portal/monitorAxios/device/metricQuery/metricQueryByDevResourceId'
    headers = {
        'content-type': 'application/json;charset=UTF-8',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.67',
        'token': token
    }
    pool_dic = {
        '6180578038500323': '江苏',
        '6180578038500313': '甘肃',
        '6180578038500318': '河北',
        '6180578038500334': '天津',
        '6180578038500333': '山西',
        '6180578038500326': '内蒙古',
        '6180578038500314': '广西',
        '6180578038500328': '青海'
    }
    value = pool_dic.get(poolcicd)
    filename = str(value) + '.xlsx'
    i = 0
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for key, value in dic_cicd.items():
            future = executor.submit(
                process_cicd_entry,
                key,
                value,
                b_time,
                e_time,
                url,
                filename,
                headers,
                count,
                choose_type
            )
            futures.append(future)

        for future in concurrent.futures.as_completed(futures):
            i += 1
            print("progress: {}/{}".format(i, numbers))
            sys.stdout.flush()
            result = future.result()
            print(result)

    # ... （保留文件关闭部分）...


def process_cicd_entry(key, value, b_time, e_time, url, filename, headers, count, choose_type):
    def fetch_data():
        response = requests.post(url=url, headers=headers, json=json_data)
        response.raise_for_status()
        return json.loads(response.text)

    max_retries = 5
    retry_count = 0
    if choose_type == '云主机':
        json_data = {
            "page": {
                "current": 1,
                "size": 10,
                "total": 0
            },
            "query": {
                "beginTime": b_time,
                "ciCode": value,
                "dataSource": [
                    "31省"
                ],
                "0": "31省",
                "endTime": e_time,
                "devName": value,
                "devType": "VmServer",
                "maxVal": "",
                "metricNames": [
                    "cpu_usage_percent"
                ],
                "0": "cpu_usage_percent",
                "valueRange": [
                    'null',
                    'null'
                ],
                "0": "null",
                "1": "null"
            }
        }
        data_dic = {}
        while retry_count < max_retries:
            try:
                data = fetch_data()
                if data['data']['total'] < count - 10:
                    data_dic[key] = data['data']['total']
                break
            except requests.RequestException as e:
                print("Error:", e)
                retry_count += 1
                print(f"Retrying... (Attempt {retry_count}/{max_retries})")
                time.sleep(1)
        # print(data_dic)
        write_to_excel(data_dic, filename)
    else:
        json_data = {
            "page": {
                "current": 1,
                "size": 10
            },
            "query": {
                "beginTime": b_time,
                "ciCode": value,
                "dataSource": [
                    "31省"
                ],
                "0": "31省",
                "endTime": e_time,
                "maxVal": "",
                "metricNames": [
                    "cpu_usage_percent"
                ],
                "0": "cpu_usage_percent",
                "minVal": "",
                "orderBy": "",
                "timeType": "ts"
            }
        }

        data_dic = {}
        while retry_count < max_retries:
            try:
                data = fetch_data()
                if data['data']['total'] < count - 10:
                    data_dic[key] = data['data']['total']
                break
            except requests.RequestException as e:
                print("Error:", e)
                retry_count += 1
                print(f"Retrying... (Attempt {retry_count}/{max_retries})")
                time.sleep(1)
        # print(data_dic)
        write_to_excel(data_dic, filename)
    # 不再需要单独关闭文件，因为每次写入都在 `with open` 语句中完成


def write_to_excel(data_dic, target_sheet_name, filename):
    if not os.path.isfile(filename):
        # 文件不存在，创建新文件
        wb = Workbook()
        wb.active.title = '物理机'
        wb.create_sheet('云主机')
        wb.save(filename)
    else:
        # 文件已存在，打开工作簿
        wb = load_workbook(filename)

    # 清空并写入数据到指定工作表
    ws = wb[target_sheet_name]  # 获取指定工作表
    ws.delete_rows(1, ws.max_row)  # 清空工作表数据

    for row_num, (key, value) in enumerate(data_dic.items(), start=1):  # 从第1行开始写入
        ws.cell(row=row_num, column=1, value=key)
        ws.cell(row=row_num, column=2, value=value)

    wb.save(filename)


if __name__ == "__main__":
    sys.exit(main())
